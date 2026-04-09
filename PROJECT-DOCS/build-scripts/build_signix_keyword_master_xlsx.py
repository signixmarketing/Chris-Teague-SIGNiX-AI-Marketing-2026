#!/usr/bin/env python3
"""
SIGNiX Keyword Master — Excel Build Script
Generates: SIGNiX_Keyword_Master_4.6.26.xlsx
Run:  python3 "PROJECT-DOCS/build-scripts/build_signix_keyword_master_xlsx.py"
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# ── BRAND TOKENS ──────────────────────────────────────────────────────────────
GREEN       = "6da34a"
INK         = "2e3440"
BODY        = "545454"
MUTED       = "6b7280"
WHITE       = "ffffff"
CANVAS      = "f8fafb"
RULE        = "d8dee9"

# Status colors (background / text pairs)
STATUS_COLORS = {
    "Active":             ("d5e8d4", "1a5c2a"),   # green
    "Testing — Add Now":  ("fff2cc", "7d5a00"),   # amber
    "Testing Pending":    ("dae8fc", "1e4d7b"),   # blue
    "Removed — Zero Vol": ("f8cecc", "8b0000"),   # red
    "Excluded — Wrong Buyer":    ("ffe6cc", "7d3a00"),   # orange
    "Excluded — Informational":  ("f0f0f0", "444444"),   # gray
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color=BODY, bold=False, size=10, name="Calibri"):
    return Font(color=hex_color, bold=bold, size=size, name=name)

def border_thin():
    thin = Side(style="thin", color=RULE)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def border_bottom_only():
    thin = Side(style="thin", color=RULE)
    return Border(bottom=thin)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, row_num, labels, col_widths):
    """Write a styled header row."""
    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.fill = fill(INK)
        cell.font = font(WHITE, bold=True, size=10)
        cell.alignment = center()
        cell.border = border_thin()
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def apply_data_row(ws, row_num, values, status):
    """Write a data row with status-based color coding."""
    bg, fg = STATUS_COLORS.get(status, ("ffffff", BODY))
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border = border_thin()
        cell.alignment = left() if col_idx in (1, 2, 3, 12) else center()
        # Status column gets the status color; other cols get a lighter tint
        if col_idx == 4:
            cell.fill = fill(bg)
            cell.font = font(fg, bold=True, size=9)
        else:
            cell.fill = fill("fafafa")
            cell.font = font(BODY, size=9)

def section_label(ws, row_num, label, num_cols):
    """Write a full-width section divider row."""
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=num_cols)
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.fill = fill(GREEN)
    cell.font = font(WHITE, bold=True, size=10)
    cell.alignment = left()
    ws.row_dimensions[row_num].height = 20

# ── KEYWORD DATA ──────────────────────────────────────────────────────────────
# Columns: Keyword | Bucket | Sub-Bucket | Status | Avg Mo Searches
#          3-Mo Change | YoY Change | Competition Level | Competition Index
#          CPC Low | CPC High | Notes

HEADERS = [
    "Keyword",
    "Bucket",
    "Sub-Bucket / Sector",
    "Status",
    "Avg Mo Searches",
    "3-Mo Change",
    "YoY Change",
    "Competition",
    "Comp Index (0–100)",
    "CPC Low",
    "CPC High",
    "Notes / Rationale",
]

COL_WIDTHS = [38, 22, 22, 24, 16, 13, 13, 14, 18, 10, 10, 48]

# Status constants
ACTIVE      = "Active"
TEST_NOW    = "Testing — Add Now"
TEST_PEND   = "Testing Pending"
REMOVED     = "Removed — Zero Vol"
EX_BUYER    = "Excluded — Wrong Buyer"
EX_INFO     = "Excluded — Informational"

# fmt: Keyword | Bucket | Sub-Bucket | Status | Vol | 3Mo | YoY | CompLvl | CompIdx | CPCLow | CPCHigh | Notes

KEYWORDS = [
    # ── SECTION: ACTIVE ───────────────────────────────────────────────────────
    ("__SECTION__", "ACTIVE — Validated, in campaign"),

    ("beneficiary change form",
     "B3 — Wealth Management", "Wealth / Advisor",
     ACTIVE, 390, "+23%", "+50%", "Low", 2, "$1.02", "$3.93",
     "Best ROI of all keywords tested. Anchor term. Already in campaign. Low comp, decent volume, growing YoY."),

    ("hipaa compliant e-signature",
     "B4 — Industry / Use Case", "Healthcare",
     ACTIVE, 210, "+191%", "+52%", "Low", 11, "$11.07", "$53.94",
     "Fastest growing keyword in the set. Already in campaign. High CPC but strong buyer intent. Anchor term."),

    ("legal electronic signature",
     "B1 — Compliance / Legal", "Legal Validity",
     ACTIVE, 390, "+24%", "-46%", "Low", 11, "$3.97", "$22.67",
     "Solid volume, low comp. Already in account. YoY declining — monitor quarterly. Keep for now."),

    ("docusign competitors",
     "B5 — Competitor Displacement", "Competitor",
     ACTIVE, 1000, "0%", "-32%", "Medium", 63, "$13.00", "$68.14",
     "Best-intent competitor term. Already in campaign. YoY declining but still best volume in displacement bucket."),

    ("docusign alternative",
     "B5 — Competitor Displacement", "Competitor",
     ACTIVE, 390, "0%", "-33%", "High", 69, "$10.57", "$39.65",
     "Already in account. Declining. Keep alongside 'docusign competitors'. Watch for further YoY decline."),

    # ── SECTION: TESTING — ADD NOW ────────────────────────────────────────────
    ("__SECTION__", "TESTING — Add Now (validated in Keyword Planner, recommended for next campaign build)"),

    ("are electronic signatures legally binding",
     "B1 — Compliance / Legal", "Legal Validity",
     TEST_NOW, 260, "+24%", "+53%", "Low", 3, "$0.95", "$5.75",
     "Growing fast. Low CPC. Buyers researching before switching. Strong 'document must stand up to authority' match."),

    ("signatures on contracts",
     "B1 — Compliance / Legal", "Legal Validity",
     TEST_NOW, 390, "0%", "0%", "Low", 29, "$4.52", "$17.14",
     "Solid volume, stable. Broad but low comp. Good top-of-funnel for legal sector. Test with tightly matched ad copy."),

    ("are digital signatures legally binding",
     "B1 — Compliance / Legal", "Legal Validity",
     TEST_NOW, 140, "+57%", "-21%", "Low", 5, "$2.00", "$8.35",
     "Fastest 3-mo growth in legal validity bucket. PKI angle maps cleanly. Test alongside electronic sig variant."),

    ("legal document signing",
     "B1 — Compliance / Legal", "Legal Validity",
     TEST_NOW, 140, "+22%", "-21%", "Low", 21, "$2.74", "$16.87",
     "Workflow-specific. Good intent. Declining YoY but growing 3-mo. Use for legal / debt collection sector."),

    ("electronic signature verification statement example",
     "B1 — Compliance / Legal", "Legal Validity",
     TEST_NOW, 110, "n/a", "+1500%", "Low", 0, "—", "—",
     "Near-zero comp. Surging YoY (new search behavior). Likely driven by court filing requirements. Test immediately."),

    ("e notary platforms",
     "B6 — RON Institutional", "RON",
     TEST_NOW, 140, "—", "+27%", "Low", 26, "$1.62", "$8.06",
     "Low CPC, growing. Better buyer language than 'RON platform.' Use for notary and financial institution audiences."),

    ("hipaa electronic signature",
     "B4 — Industry / Use Case", "Healthcare",
     TEST_NOW, 70, "0%", "-67%", "Medium", 36, "$8.18", "$27.00",
     "Companion to 'hipaa compliant e-signature.' Declining YoY but solid healthcare buyer intent. Test as second keyword."),

    ("e signature consent form",
     "B4 — Industry / Use Case", "Healthcare / Legal",
     TEST_NOW, 40, "+67%", "+25%", "Low", 20, "$3.26", "$15.65",
     "Growing fast. Consent form = authority-facing doc. Strong match for healthcare + debt collection sectors."),

    ("e signature consent",
     "B4 — Industry / Use Case", "Healthcare / Legal",
     TEST_NOW, 70, "0%", "0%", "Low", 24, "$3.89", "$24.23",
     "Stable. Pairs well with 'e signature consent form.' Low CPC. Good for patient and debtor consent workflows."),

    ("digital signatures pki",
     "B2 — Authentication / Fraud", "Technical / PKI",
     TEST_NOW, 70, "+40%", "-36%", "Low", 10, "$2.96", "$13.11",
     "Low CPC, low comp. Technical buyer. Differentiator: SIGNiX is PKI-based vs image-based. Good for partner/dev audience."),

    ("electronic notarization platform",
     "B6 — RON Institutional", "RON",
     TEST_NOW, 30, "growing", "new", "Low", 3, "—", "—",
     "Emerging term. Near-zero comp. RON is a proven SIGNiX lead channel — capture early before competition builds."),

    ("power of attorney e-signature",
     "B3 — Wealth Management", "Wealth / Advisor",
     TEST_NOW, 30, "flat", "+50%", "Low", "—", "—", "—",
     "Growing YoY. POA = high-stakes authority doc. Perfect use case for PKI + audit trail. Add to wealth mgmt ad group."),

    # ── SECTION: TESTING PENDING ──────────────────────────────────────────────
    ("__SECTION__", "TESTING PENDING — In plan but not yet run in Google Keyword Planner"),

    # Auth / Fraud (CEO priority)
    ("electronic signature fraud prevention banking",
     "B2 — Authentication / Fraud", "Auth / Fraud — Banking",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "CEO priority term. Fraud angle opens doors per CEO. Run in Keyword Planner first — may return zero volume."),

    ("identity verification electronic signature",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "CEO priority: ID Verify. Corrected word order vs. removed 'e signature identity verification.' Test now."),

    ("e-signature with identity verification",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Companion to above. Test both; keep whichever has volume."),

    ("verified electronic signature",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Shorter form. May have broader buyer reach. CEO / ID Verify priority."),

    ("electronic signature signer authentication",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Long-tail. Likely low volume but high intent. Run before excluding."),

    ("two-factor authentication electronic signature",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "2FA is a known concept. May have search volume. Run test."),

    ("biometric e-signature financial services",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Biometric = strong AI-fraud counter-narrative. Run test. Likely long-tail with low volume."),

    ("signer authentication platform",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Platform-level buyer language. Test alongside identity verification terms."),

    # Wealth Management
    ("POA document signing platform",
     "B3 — Wealth Management", "Wealth / Advisor",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Companion to 'power of attorney e-signature.' Likely low volume — run before committing budget."),

    ("investment account opening signature",
     "B3 — Wealth Management", "Wealth / Advisor",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "New account opening is a core SIGNiX use case. Test — may surface in brokerage / advisor searches."),

    ("401k distribution electronic signature",
     "B3 — Wealth Management", "Wealth / Advisor",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "'ira distribution e sign' was zero. Re-test 401k variant — different buyer language."),

    ("wealth management document signing",
     "B3 — Wealth Management", "Wealth / Advisor",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Broader phrase. May have more volume than specific variants. Test as category anchor."),

    ("financial advisor e-signature compliance",
     "B3 — Wealth Management", "Wealth / Advisor",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Compliance angle + advisor persona. Strong intent if volume exists."),

    ("e-signature for investment advisors",
     "B3 — Wealth Management", "Wealth / Advisor",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Direct persona targeting. Test — may have niche volume from RIA community."),

    # Banking / FI
    ("e-signature for credit unions",
     "B4 — Industry / Use Case", "Banking / FI",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Credit unions = strong TruStage channel. Test before activating. May mirror 'e signature for banks' (20/mo)."),

    ("loan document electronic signature",
     "B4 — Industry / Use Case", "Banking / FI",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Core FI use case. Loan docs are primary TruStage workflow. Test — if volume exists, activate immediately."),

    ("wire transfer authorization form",
     "B4 — Industry / Use Case", "Banking / FI",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Fraud-angle keyword. Wire auth = high-risk transaction needing authentication. CEO priority use case."),

    ("wire transfer form electronic signature",
     "B4 — Industry / Use Case", "Banking / FI",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Companion to above. Test both word orders."),

    ("ACH authorization e-signature",
     "B4 — Industry / Use Case", "Banking / FI",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "ACH auth = consent + authentication = SIGNiX sweet spot. Run test."),

    ("new account opening electronic signature",
     "B4 — Industry / Use Case", "Banking / FI",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Account opening is a 'Width' use case (beyond loan docs). Test with depth/width strategy in mind."),

    ("electronic signature NCUA compliant",
     "B4 — Industry / Use Case", "Banking / FI",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Regulatory compliance keyword for credit unions. Likely low volume. Run before excluding."),

    # Healthcare
    ("patient consent electronic signature",
     "B4 — Industry / Use Case", "Healthcare",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Patient consent = authority-facing doc. Test alongside HIPAA keywords."),

    ("telehealth consent form signature",
     "B4 — Industry / Use Case", "Healthcare",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Telehealth grew post-COVID. Consent forms are a SIGNiX use case. May have decent volume."),

    ("medical records release e-signature",
     "B4 — Industry / Use Case", "Healthcare",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "HIPAA-adjacent. Medical records release = official, authority-facing. Run test."),

    # Legal / Debt Collection (new sector)
    ("FDCPA compliant e-signature",
     "B4 — Industry / Use Case", "Legal / Debt Collection",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "New sector. FDCPA = collection agencies must comply with strict consent rules. SIGNiX is a natural fit."),

    ("debt settlement agreement electronic signature",
     "B4 — Industry / Use Case", "Legal / Debt Collection",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Settlement agreements = signed contracts under authority. Ideal SIGNiX use case."),

    ("debt collection consent form",
     "B4 — Industry / Use Case", "Legal / Debt Collection",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Consent form = authentication need. Maps directly to ID Verify. Test as priority for new sector."),

    ("legal document e-signature platform",
     "B4 — Industry / Use Case", "Legal / Debt Collection",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Broad legal platform term. May compete with Clio/DocuSign in legal. Test before activating."),

    ("collection agency e-signature",
     "B4 — Industry / Use Case", "Legal / Debt Collection",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Niche but targeted. Collection agencies need compliant consent + signatures. Run test."),

    ("attorney e-signature compliance",
     "B4 — Industry / Use Case", "Legal / Debt Collection",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Attorney persona. Compliance angle. Likely low volume but high intent. Test before excluding."),

    # Compliance (SEC / FINRA) — on hold
    ("SEC Rule 17a-4 compliant e-signature",
     "B1 — Compliance / Legal", "Regulatory — SEC",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "In plan. Similar terms tested returned zero volume. Run exact phrase — SEC audience may search differently."),

    ("FINRA electronic signature audit trail",
     "B1 — Compliance / Legal", "Regulatory — FINRA",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "'finra digital signature' = zero volume. Try audit trail variant before removing entire FINRA bucket."),

    ("non-repudiation e-signature financial services",
     "B1 — Compliance / Legal", "Regulatory — SEC",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Non-repudiation is SIGNiX's core PKI advantage. Test — may surface among compliance officers."),

    ("digital signature SEC compliance",
     "B1 — Compliance / Legal", "Regulatory — SEC",
     TEST_PEND, "—", "—", "—", "—", "—", "—", "—",
     "Slight variation on 'sec 17a 4 e signature' (zero vol). Run test — compliance buyer language may differ."),

    # ── SECTION: REMOVED — ZERO VOLUME ────────────────────────────────────────
    ("__SECTION__", "REMOVED — Zero Volume (tested in Keyword Planner, confirmed dead)"),

    ("e signature for banks",
     "B4 — Industry / Use Case", "Banking / FI",
     REMOVED, 20, "0%", "-75%", "Low", 10, "—", "—",
     "Dying term. Volume too low and falling -75% YoY. Do not reactivate."),

    ("digital signature mortgage docs",
     "B4 — Industry / Use Case", "Mortgage (removed sector)",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Sector removed: mortgage replaced by Legal / Debt Collection."),

    ("ron platform financial institutions",
     "B6 — RON Institutional", "RON",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Jargon. Buyers don't search 'RON platform.' Use 'e notary platforms' instead."),

    ("e signature rollover forms",
     "B3 — Wealth Management", "Wealth / Advisor",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Too product-specific. Not how buyers search. Do not retest."),

    ("ira distribution e sign",
     "B3 — Wealth Management", "Wealth / Advisor",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Not buyer language. Re-test as '401k distribution electronic signature' (different phrase)."),

    ("kba electronic signature",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. KBA is internal jargon. Buyers search for 'identity verification' not 'KBA.' Do not retest."),

    ("e signature identity verification",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Wrong word order. Re-test as 'identity verification electronic signature' (Testing Pending)."),

    ("ai fraud e signature",
     "B2 — Authentication / Fraud", "Auth / Fraud",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume now. Category is emerging. Re-test Q3 2026 as AI fraud news grows."),

    ("ron platform compliance",
     "B6 — RON Institutional", "RON",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Not buyer language. Use 'electronic notarization platform' instead."),

    ("remote online notary for banks",
     "B6 — RON Institutional", "RON",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Too narrow. 'e notary platforms' performs better."),

    ("docusign alternative for banks",
     "B5 — Competitor Displacement", "Competitor",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Industry modifier kills search. Use 'docusign alternative' without modifier."),

    ("adobe sign alternative compliance",
     "B5 — Competitor Displacement", "Competitor",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Not how buyers search. Adobe modifier + compliance = too narrow."),

    ("sec 17a 4 e signature",
     "B1 — Compliance / Legal", "Regulatory — SEC",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Regulatory jargon only. Try 'SEC Rule 17a-4 compliant e-signature' as a testing-pending variant."),

    ("finra digital signature",
     "B1 — Compliance / Legal", "Regulatory — FINRA",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Confirmed. Try 'FINRA electronic signature audit trail' as alternate phrasing (Testing Pending)."),

    ("electronic recordkeeping broker dealer",
     "B1 — Compliance / Legal", "Regulatory — FINRA",
     REMOVED, 0, "—", "—", "—", "—", "—", "—",
     "Zero volume. Compliance / IT team language, not buyer search language."),

    # ── SECTION: EXCLUDED — WRONG BUYER ───────────────────────────────────────
    ("__SECTION__", "EXCLUDED — Wrong Buyer (volume exists but attracts price-sensitive SMB, not B2B enterprise)"),

    ("free docusign alternatives",
     "B5 — Competitor Displacement", "Competitor",
     EX_BUYER, 1900, "+46%", "0%", "High", 72, "$2.88", "$14.18",
     "Highest volume competitor term. But 'free' = SMB price shopper. Negative ROI. Negative keyword this phrase."),

    ("free alternative to docusign",
     "B5 — Competitor Displacement", "Competitor",
     EX_BUYER, 170, "+21%", "-19%", "High", "—", "—", "—",
     "Same issue as above. 'Free' signals price-sensitive buyer. Exclude as negative keyword."),

    ("cheaper alternative to docusign",
     "B5 — Competitor Displacement", "Competitor",
     EX_BUYER, 90, "+40%", "-22%", "High", "—", "—", "—",
     "'Cheaper' = wrong buyer. Add to negative keyword list."),

    ("cheapest docusign alternative",
     "B5 — Competitor Displacement", "Competitor",
     EX_BUYER, 70, "0%", "-55%", "High", "—", "—", "—",
     "Wrong buyer and declining. Exclude. Not our customer."),

    # ── SECTION: EXCLUDED — INFORMATIONAL ─────────────────────────────────────
    ("__SECTION__", "EXCLUDED — Informational (researcher / consumer intent, not B2B buying signal)"),

    ("is a digital signature legally binding",
     "B1 — Compliance / Legal", "Legal Research",
     EX_INFO, 170, "—", "—", "Low", "—", "—", "—",
     "Informational. Consumer / student research, not a buyer. Could test with strong negative keywords."),

    ("are e signatures legally binding",
     "B1 — Compliance / Legal", "Legal Research",
     EX_INFO, 140, "—", "—", "Low", "—", "—", "—",
     "Informational. Nearly identical to above. Do not activate without negative keyword controls."),

    ("is an electronic signature legally binding",
     "B1 — Compliance / Legal", "Legal Research",
     EX_INFO, 140, "—", "—", "Low", "—", "—", "—",
     "Informational variant. Same issue as above."),

    ("electronic signature legal requirements",
     "B1 — Compliance / Legal", "Legal Research",
     EX_INFO, 70, "—", "-71%", "Low", "—", "—", "—",
     "Informational and declining -71% YoY. Do not activate."),

    ("best docusign alternatives",
     "B5 — Competitor Displacement", "Competitor Research",
     EX_INFO, 110, "—", "-18%", "—", "—", "—", "—",
     "'Best' = research phase, not purchase ready. Declining. Do not activate."),

    ("similar to docusign",
     "B5 — Competitor Displacement", "Competitor Research",
     EX_INFO, 170, "—", "-81%", "—", "—", "—", "—",
     "Research intent. Steep -81% YoY decline. Do not activate."),

    ("docusign alternatives for small business",
     "B5 — Competitor Displacement", "Competitor Research",
     EX_INFO, 50, "—", "-40%", "—", "—", "—", "—",
     "Wrong segment (SMB). Declining. Not our buyer."),
]

# ── BUDGET ALLOCATION DATA ────────────────────────────────────────────────────
BUDGET_HEADERS = [
    "Bucket", "Monthly Budget", "3-Month Total", "% of Total",
    "Primary Keywords", "Phase 1 Status", "Notes",
]

BUDGET_COL_WIDTHS = [28, 16, 16, 13, 50, 20, 40]

BUDGET_ROWS = [
    ("B2 — Authentication / ID Fraud", "$600", "$1,800", "20%",
     "identity verification e-sig, e-sig fraud prevention banking, signer authentication platform",
     "Testing Pending",
     "CEO priority: ID Verify. Run Keyword Planner tests Q2 2026 before full activation."),
    ("B4 — Healthcare / HIPAA", "$500", "$1,500", "17%",
     "hipaa compliant e-signature (active), hipaa electronic signature, patient consent e-sig",
     "Active + Testing",
     "Anchor term already in campaign. Add HIPAA electronic sig + patient consent after testing."),
    ("B3 — Wealth Management", "$500", "$1,500", "17%",
     "beneficiary change form (active), power of attorney e-sig, wealth mgmt document signing",
     "Active + Testing",
     "Anchor term (best ROI). Add POA and 401k terms after Keyword Planner validation."),
    ("B5 — Competitor Displacement", "$400", "$1,200", "13%",
     "docusign competitors (active), docusign alternative (active)",
     "Active",
     "Both terms already in campaign. Monitor YoY decline. Do NOT activate 'free' variants."),
    ("B1 — Legal Validity / Authority", "$300", "$900", "10%",
     "are electronic signatures legally binding, legal document signing, signatures on contracts",
     "Testing — Add Now",
     "New angle from CPO insight: document must stand up to authority. Low CPC. Add in next build."),
    ("B4 — Legal / Debt Collection", "$300", "$900", "10%",
     "FDCPA compliant e-sig, debt collection consent form, debt settlement agreement e-sig",
     "Testing Pending",
     "New sector replacing mortgage. Run Keyword Planner tests first. Full activation Q2 2026."),
    ("B6 — RON Institutional", "$200", "$600", "7%",
     "e notary platforms, electronic notarization platform",
     "Testing — Add Now",
     "RON = proven SIGNiX lead channel. Both terms validated. Add now."),
    ("B4 — Banking / Financial Institutions", "$0", "$0", "0%",
     "wire transfer authorization form, loan document e-sig, ACH authorization e-sig",
     "Testing Pending",
     "On hold. Run Keyword Planner before activating. Credit unions = TruStage overlap risk."),
    ("B1 — Compliance / Regulatory (SEC/FINRA)", "$0", "$0", "0%",
     "SEC 17a-4 compliant e-sig, FINRA electronic sig audit trail",
     "Testing Pending",
     "On hold. All current terms returned zero volume. Re-test alternate phrases Q2 2026."),
    ("Reserve / Testing", "$200", "$600", "7%",
     "New keyword validation budget",
     "Rolling",
     "Validate any new term here before moving full budget. Pull from this bucket for test runs."),
    ("TOTAL", "$3,000", "$9,000", "100%", "—", "—", "—"),
]

# ── BUILD WORKBOOK ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══ SHEET 1: KEYWORD MASTER ══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Keyword Master"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

# Title row
ws1.row_dimensions[1].height = 32
ws1.merge_cells("A1:L1")
title_cell = ws1["A1"]
title_cell.value = "SIGNiX Paid Media — Keyword Master   |   Rev 3  ·  April 6, 2026"
title_cell.fill = fill(INK)
title_cell.font = Font(color=WHITE, bold=True, size=13, name="Calibri")
title_cell.alignment = Alignment(horizontal="left", vertical="center")

# Header row
apply_header_row(ws1, 2, HEADERS, COL_WIDTHS)
ws1.row_dimensions[2].height = 36

# Data rows
current_row = 3
num_cols = len(HEADERS)

for item in KEYWORDS:
    if item[0] == "__SECTION__":
        section_label(ws1, current_row, "  " + item[1], num_cols)
        current_row += 1
    else:
        (keyword, bucket, sub, status,
         vol, chg3, yoy, comp_lvl, comp_idx, cpc_lo, cpc_hi, notes) = item
        values = [keyword, bucket, sub, status,
                  vol, chg3, yoy, comp_lvl, comp_idx, cpc_lo, cpc_hi, notes]
        apply_data_row(ws1, current_row, values, status)
        ws1.row_dimensions[current_row].height = 30
        current_row += 1

# Auto-filter on header row
ws1.auto_filter.ref = f"A2:L{current_row - 1}"

# Legend block (below data)
current_row += 1
ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
legend_hdr = ws1.cell(row=current_row, column=1, value="STATUS LEGEND")
legend_hdr.fill = fill(INK)
legend_hdr.font = font(WHITE, bold=True)
legend_hdr.alignment = left()
current_row += 1

for status, (bg, fg) in STATUS_COLORS.items():
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    c = ws1.cell(row=current_row, column=1, value=status)
    c.fill = fill(bg)
    c.font = Font(color=fg, bold=True, size=9, name="Calibri")
    c.alignment = center()
    current_row += 1

# ══ SHEET 2: BUDGET ALLOCATION ════════════════════════════════════════════════
ws2 = wb.create_sheet("Budget Allocation")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

# Title
ws2.row_dimensions[1].height = 32
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "SIGNiX Paid Media — Budget Allocation   |   $3,000 / Month   ·   April 2026"
t2.fill = fill(INK)
t2.font = Font(color=WHITE, bold=True, size=13, name="Calibri")
t2.alignment = Alignment(horizontal="left", vertical="center")

# Header
apply_header_row(ws2, 2, BUDGET_HEADERS, BUDGET_COL_WIDTHS)
ws2.row_dimensions[2].height = 36

for row_idx, row in enumerate(BUDGET_ROWS, start=3):
    is_total = row[0] == "TOTAL"
    ws2.row_dimensions[row_idx].height = 40
    for col_idx, val in enumerate(row, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border_thin()
        cell.alignment = left() if col_idx in (1, 5, 6, 7) else center()
        if is_total:
            cell.fill = fill(INK)
            cell.font = font(WHITE, bold=True, size=10)
        elif row[1] == "$0":
            cell.fill = fill("f0f0f0")
            cell.font = font(MUTED, size=9)
        else:
            cell.fill = fill("fafafa")
            cell.font = font(BODY, size=9)

# ══ SHEET 3: NEGATIVE KEYWORDS ═══════════════════════════════════════════════
ws3 = wb.create_sheet("Negative Keywords")
ws3.sheet_view.showGridLines = False

ws3.row_dimensions[1].height = 32
ws3.merge_cells("A1:C1")
t3 = ws3["A1"]
t3.value = "SIGNiX — Negative Keyword List   |   Block these to protect budget"
t3.fill = fill(INK)
t3.font = Font(color=WHITE, bold=True, size=13, name="Calibri")
t3.alignment = Alignment(horizontal="left", vertical="center")

neg_headers = ["Negative Keyword", "Match Type", "Reason"]
neg_widths = [38, 16, 60]
apply_header_row(ws3, 2, neg_headers, neg_widths)
ws3.row_dimensions[2].height = 36

NEGATIVE_KEYWORDS = [
    ("free",                "Broad",    "Attracts price-sensitive SMB searchers. Not our buyer."),
    ("free docusign alternatives", "Phrase", "1,900 searches/mo — all wrong buyer. Must block."),
    ("free alternative to docusign", "Phrase", "Same as above."),
    ("cheaper",             "Broad",    "Price-motivated. Not our buyer persona."),
    ("cheapest",            "Broad",    "Same."),
    ("small business",      "Phrase",   "SMB-focused searches; SIGNiX sells to mid-market and enterprise."),
    ("free trial",          "Phrase",   "Freeware intent. Negative ROI."),
    ("open source",         "Phrase",   "Not our product category."),
    ("personal",            "Broad",    "Consumer use case, not B2B."),
    ("individual",          "Broad",    "Same as personal."),
    ("student",             "Broad",    "Academic, not commercial buyer."),
    ("free pdf signer",     "Phrase",   "Consumer PDF tool searcher."),
    ("how to sign a pdf",   "Phrase",   "Informational / consumer."),
    ("docusign pricing",    "Phrase",   "Comparison shopping, not necessarily our buyer."),
    ("adobe sign pricing",  "Phrase",   "Same as above."),
    ("what is electronic signature", "Phrase", "Top-of-funnel educational. Not purchase intent."),
]

for row_idx, (kw, match, reason) in enumerate(NEGATIVE_KEYWORDS, start=3):
    ws3.row_dimensions[row_idx].height = 24
    for col_idx, val in enumerate([kw, match, reason], start=1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill("fafafa")
        cell.font = font(BODY, size=9)
        cell.border = border_thin()
        cell.alignment = left()

# ── SAVE ──────────────────────────────────────────────────────────────────────
OUT = ("/Users/chris/Desktop/AI Summit 2026 Q2/"
       "proj-template-and-lease-SIGNiX-app/PROJECT-DOCS/"
       "SIGNiX_Keyword_Master_4.6.26.xlsx")

wb.save(OUT)
print(f"Saved → {OUT}")
